#!/usr/bin/env python3
"""Inspect a local Fundus-AVSeg dataset checkout.

The script validates the expected folder layout, image/mask pairing, official
split files, image sizes, and RGB annotation colors.
"""

from __future__ import annotations

import argparse
from collections import Counter
from pathlib import Path
from typing import Iterable


PROJECT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_ROOT = PROJECT_DIR / "data" / "Fundus-AVSeg"

CLASS_COLORS = {
    (0, 0, 0): "background",
    (255, 0, 0): "artery",
    (0, 0, 255): "vein",
    (0, 255, 0): "crossing",
    (255, 255, 255): "uncertain",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Inspect the local Fundus-AVSeg dataset structure and masks."
    )
    parser.add_argument(
        "--root",
        type=Path,
        default=DEFAULT_ROOT,
        help=f"Dataset root directory. Default: {DEFAULT_ROOT}",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Limit image/mask pixel inspection to the first N samples. 0 means all.",
    )
    parser.add_argument(
        "--max-unknown-colors",
        type=int,
        default=20,
        help="Maximum number of unknown RGB colors to print.",
    )
    return parser.parse_args()


def read_split(path: Path) -> list[str]:
    if not path.exists():
        return []
    names: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line:
            names.append(line)
    return names


def list_png_names(path: Path) -> set[str]:
    if not path.exists():
        return set()
    return {item.name for item in path.glob("*.png") if item.is_file()}


def print_counter(title: str, counter: Counter) -> None:
    print(f"\n{title}")
    if not counter:
        print("  none")
        return
    for key, count in counter.most_common():
        print(f"  {key}: {count}")


def summarize_missing(title: str, names: Iterable[str], limit: int = 10) -> bool:
    names = sorted(names)
    if not names:
        return False
    print(f"\n{title}: {len(names)}")
    for name in names[:limit]:
        print(f"  {name}")
    if len(names) > limit:
        print(f"  ... {len(names) - limit} more")
    return True


def inspect_metadata(path: Path) -> bool:
    if not path.exists():
        print("\nmetadata.xlsx: missing")
        return True

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("\nmetadata.xlsx: exists (install openpyxl to inspect sheet contents)")
        return False

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    print("\nmetadata.xlsx")
    print(f"  sheet: {sheet.title}")
    print(f"  rows: {sheet.max_row}")
    print(f"  columns: {sheet.max_column}")
    print(f"  headers: {headers}")
    workbook.close()
    return False


def inspect_pixels(
    image_dir: Path,
    mask_dir: Path,
    names: list[str],
    limit: int,
    max_unknown_colors: int,
) -> bool:
    try:
        from PIL import Image
    except ImportError as exc:
        raise SystemExit("Pillow is required to inspect Fundus-AVSeg images.") from exc

    if limit > 0:
        names = names[:limit]

    image_sizes: Counter[tuple[int, int]] = Counter()
    mask_sizes: Counter[tuple[int, int]] = Counter()
    class_pixels: Counter[str] = Counter()
    unknown_colors: Counter[tuple[int, int, int]] = Counter()
    bad = False
    mismatched_sizes: list[str] = []
    too_many_colors: list[str] = []

    for name in names:
        image_path = image_dir / name
        mask_path = mask_dir / name
        if not image_path.exists() or not mask_path.exists():
            continue

        with Image.open(image_path) as image:
            image_size = image.size
            image_sizes[image_size] += 1

        with Image.open(mask_path) as mask:
            rgb_mask = mask.convert("RGB")
            mask_size = rgb_mask.size
            mask_sizes[mask_size] += 1
            colors = rgb_mask.getcolors(maxcolors=rgb_mask.width * rgb_mask.height + 1)

        if image_size != mask_size:
            mismatched_sizes.append(name)
            bad = True

        if colors is None:
            too_many_colors.append(name)
            bad = True
            continue

        for count, color in colors:
            label = CLASS_COLORS.get(color)
            if label is None:
                unknown_colors[color] += count
            else:
                class_pixels[label] += count

    print(f"\nPixel-inspected samples: {len(names)}")
    print_counter("Image sizes", image_sizes)
    print_counter("Mask sizes", mask_sizes)
    print_counter("Mask class pixels", class_pixels)

    if mismatched_sizes:
        summarize_missing("Image/mask size mismatches", mismatched_sizes)
    if too_many_colors:
        summarize_missing("Masks with too many unique colors", too_many_colors)
    if unknown_colors:
        bad = True
        print(f"\nUnknown mask colors: {len(unknown_colors)}")
        for color, count in unknown_colors.most_common(max_unknown_colors):
            print(f"  {color}: {count} pixels")
        remaining = len(unknown_colors) - max_unknown_colors
        if remaining > 0:
            print(f"  ... {remaining} more")

    return bad


def main() -> int:
    args = parse_args()
    root = args.root.expanduser().resolve()
    image_dir = root / "images"
    mask_dir = root / "annotation"
    metadata_path = root / "metadata.xlsx"
    train_path = root / "training.txt"
    test_path = root / "testing.txt"

    print(f"Fundus-AVSeg root: {root}")

    bad = False
    for path in (root, image_dir, mask_dir):
        if not path.exists():
            print(f"Missing required path: {path}")
            bad = True

    if bad:
        return 1

    image_names = list_png_names(image_dir)
    mask_names = list_png_names(mask_dir)
    train_names = read_split(train_path)
    test_names = read_split(test_path)
    split_names = train_names + test_names

    print("\nFiles")
    print(f"  images: {len(image_names)}")
    print(f"  annotations: {len(mask_names)}")
    print(f"  training.txt entries: {len(train_names)}")
    print(f"  testing.txt entries: {len(test_names)}")

    bad |= summarize_missing("Images without annotation", image_names - mask_names)
    bad |= summarize_missing("Annotations without image", mask_names - image_names)

    train_set = set(train_names)
    test_set = set(test_names)
    bad |= summarize_missing("Duplicate train entries", duplicated(train_names))
    bad |= summarize_missing("Duplicate test entries", duplicated(test_names))
    bad |= summarize_missing("Train/test overlap", train_set & test_set)
    bad |= summarize_missing("Split entries missing images", set(split_names) - image_names)
    bad |= summarize_missing("Images missing from official splits", image_names - set(split_names))

    bad |= inspect_metadata(metadata_path)

    ordered_names = sorted(image_names & mask_names)
    bad |= inspect_pixels(
        image_dir=image_dir,
        mask_dir=mask_dir,
        names=ordered_names,
        limit=args.limit,
        max_unknown_colors=args.max_unknown_colors,
    )

    print("\nResult")
    if bad:
        print("  inspection failed")
        return 1
    print("  inspection passed")
    return 0


def duplicated(names: list[str]) -> set[str]:
    counts = Counter(names)
    return {name for name, count in counts.items() if count > 1}


if __name__ == "__main__":
    raise SystemExit(main())
